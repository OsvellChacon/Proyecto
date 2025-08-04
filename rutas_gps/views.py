import json
from django.shortcuts import render, get_object_or_404, redirect
from .models import Camiones, UbicacionCamion
from django.http import JsonResponse
from geopy.distance import geodesic
import openrouteservice
from .forms import VehiculoForm
from django.db.models import Q
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
import pandas as pd
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import requests

ors_client = openrouteservice.Client(key='MI_CLAVE_DE_API_DE_OPENROUTESERVICE')

# Vista para mostrar el mapa

@login_required
def mapa(request):
    camiones = Camiones.objects.all()
    ubicaciones = []

    for camion in camiones:
        ultima_ubicacion = UbicacionCamion.objects.filter(camion=camion).order_by('-timestamp').first()
        if ultima_ubicacion:
            ubicaciones.append({
                'conductor': camion.conductor.get_full_name(),  # Asegúrate de obtener el nombre completo del conductor
                'placa': camion.placa,
                'marca': camion.marca,
                'modelo': camion.modelo,
                'lat': ultima_ubicacion.latitude,
                'lng': ultima_ubicacion.longitude
            })

    return render(request, 'rutas_gps/index.html', {'camiones': camiones, 'ubicaciones': ubicaciones})


def calcular_duracion(request):
    if request.method == 'POST':
        lat_inicio = float(request.POST['lat_inicio'])
        lng_inicio = float(request.POST['lng_inicio'])
        lat_final = float(request.POST['lat_final'])
        lng_final = float(request.POST['lng_final'])
        camion_id = request.POST['camion_id']

        # Calcular la distancia usando geopy (en kilómetros)
        inicio = (lat_inicio, lng_inicio)
        final = (lat_final, lng_final)
        distancia = geodesic(inicio, final).kilometers

        # Obtener el camión y su capacidad del tanque
        camion = Camiones.objects.get(id=camion_id)
        capacidad_tanque = camion.capacidad_tanque

        # Suponiendo un rendimiento promedio de 10 km/l (puedes ajustar este valor según el vehículo)
        rendimiento_promedio = 10
        consumo_gasolina = distancia / rendimiento_promedio  # en litros

        # Verificar si la capacidad del tanque es suficiente para el viaje
        if consumo_gasolina > capacidad_tanque:
            return JsonResponse({'status': 'error', 'message': 'La capacidad del tanque no es suficiente para este viaje.'})

        # Obtener la duración del viaje usando la API de OpenRouteService
        directions_result = ors_client.directions(
            coordinates=[[lng_inicio, lat_inicio], [lng_final, lat_final]],
            profile='driving-car',
            format='geojson',
            validate=True
        )

        # Extraer la duración del viaje en segundos
        duracion_segundos = directions_result['features'][0]['properties']['segments'][0]['duration']
        duracion_minutos = duracion_segundos / 60

        return JsonResponse({'duracion': duracion_minutos, 'distancia': distancia, 'consumo_gasolina': consumo_gasolina})

    return render(request, 'rutas_gps/index.html')

# Vista para obtener las ubicaciones de los camiones en formato JSON (para la API)
def ubicaciones_camiones(request):
    camiones = Camiones.objects.all()
    data = []

    for camion in camiones:
        ultima_ubicacion = UbicacionCamion.objects.filter(camion=camion).order_by('-timestamp').first()
        if ultima_ubicacion:
            data.append({
                'conductor': camion.conductor,  # Asegúrate de obtener el nombre completo del conductor
                'placa': camion.placa,
                'marca': camion.marca,
                'modelo': camion.modelo,
                'lat': ultima_ubicacion.latitude,
                'lng': ultima_ubicacion.longitude
            })

    return JsonResponse(data, safe=False)

def guardar_ruta(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        camion_id = data['camion_id']
        ruta = data['ruta']
        gasolina_inicio = float(data.get('gasolina_inicio', 0))

        camion = Camiones.objects.get(id=camion_id)
        capacidad_tanque = camion.capacidad_tanque

        # Validación de gasolina inicial
        if gasolina_inicio > capacidad_tanque:
            return JsonResponse({'status': 'error', 'message': 'La gasolina inicial no puede ser mayor que la capacidad del tanque.'})

        # Calcular distancia total (en km)
        puntos = ruta
        distancia_total = 0
        for i in range(1, len(puntos)):
            lat1, lng1 = puntos[i-1]['lat'], puntos[i-1]['lng']
            lat2, lng2 = puntos[i]['lat'], puntos[i]['lng']
            distancia_total += geodesic((lat1, lng1), (lat2, lng2)).kilometers

        rendimiento_promedio = 10  # km/litro
        consumo = distancia_total / rendimiento_promedio

        gasolina_fin = max(gasolina_inicio - consumo, 0)
        porcentaje_ahorrado = (gasolina_fin / gasolina_inicio * 100) if gasolina_inicio > 0 else 0

        # Calcular duración del viaje (simulación: velocidad promedio 60km/h)
        velocidad_promedio = 60  # km/h
        duracion_viaje = (distancia_total / velocidad_promedio) * 60 if velocidad_promedio > 0 else 0  # minutos

        UbicacionCamion.objects.create(
            camion=camion,
            latitude=ruta[0]['lat'],
            longitude=ruta[0]['lng'],
            ruta=ruta,
            viaje_completado=True,
            gasolina_inicio=gasolina_inicio,
            gasolina_fin=gasolina_fin,
            porcentaje_ahorrado=porcentaje_ahorrado,
            duracion_viaje=duracion_viaje
        )

        return JsonResponse({'status': 'success', 'message': 'Ruta guardada correctamente'})

    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)


def vehiculos_list(request):
    query = request.GET.get('q', '')
    users = Camiones.objects.order_by('-id')

    if query:
        users = users.filter(
            Q(conductor__first_name__icontains=query) |
            Q(conductor__last_name__icontains=query) |
            Q(placa__icontains=query) |
            Q(marca__icontains=query) |
            Q(modelo__icontains=query)
        )

        if not users.exists():
            messages.warning(request, 'No se encontraron vehiculos con los criterios de búsqueda proporcionados.')

    # Agregar paginación
    page = request.GET.get('page', 1)
    paginator = Paginator(users, 5)  # Muestra 5 usuarios por página

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
        
    megumi = {
        'users': users,
        'page_title': 'Oseed | Vehiculos'
    }

    return render(request, 'vehiculos/vehiculos.html', megumi)

@login_required
def detalles_vehiculos(request, user_id):
    user = get_object_or_404(Camiones, pk=user_id)
    viajes = UbicacionCamion.objects.filter(camion=user, viaje_completado=True).order_by('-fecha_viaje')

    # Estadísticas avanzadas
    total_km = 0
    total_gasolina = 0
    total_duracion = 0
    total_viajes = viajes.count()
    eficiencia = 0
    alertas = []

    def geocode(lat, lng):
        try:
            url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lng}&zoom=16&addressdetails=1"
            resp = requests.get(url, headers={'User-Agent': 'OseedApp'})
            if resp.status_code == 200:
                data = resp.json()
                return data.get('display_name', '')
        except Exception:
            return ""
        return ""

    viajes_info = []
    for v in viajes:
        ruta = v.ruta if v.ruta else []
        km_viaje = 0
        for i in range(1, len(ruta)):
            lat1, lng1 = ruta[i-1]['lat'], ruta[i-1]['lng']
            lat2, lng2 = ruta[i]['lat'], ruta[i]['lng']
            km_viaje += geodesic((lat1, lng1), (lat2, lng2)).kilometers
        total_km += km_viaje
        consumo_viaje = (v.gasolina_inicio - v.gasolina_fin)
        total_gasolina += consumo_viaje
        total_duracion += v.duracion_viaje

        # Alertas por viaje
        if v.gasolina_fin < (user.capacidad_tanque * 0.1):
            alertas.append(f"¡Gasolina baja en viaje del {v.fecha_viaje.strftime('%d/%m/%Y %H:%M')}! ({v.gasolina_fin:.2f} L)")
        if km_viaje > 500:
            alertas.append(f"¡Viaje largo detectado ({km_viaje:.1f} km) el {v.fecha_viaje.strftime('%d/%m/%Y %H:%M')}!")
        if km_viaje > 0 and consumo_viaje > 0 and (km_viaje / consumo_viaje) < 2:
            alertas.append(f"¡Eficiencia anómala en viaje del {v.fecha_viaje.strftime('%d/%m/%Y %H:%M')}! ({km_viaje / consumo_viaje:.2f} km/l)")

        # Geocodificación de inicio y fin
        direccion_inicio = ""
        direccion_fin = ""
        if ruta:
            direccion_inicio = geocode(ruta[0]['lat'], ruta[0]['lng'])
            direccion_fin = geocode(ruta[-1]['lat'], ruta[-1]['lng'])
        viajes_info.append({
            'obj': v,
            'direccion_inicio': direccion_inicio,
            'direccion_fin': direccion_fin,
            'km_viaje': round(km_viaje, 2)
        })

    promedio_duracion = total_duracion / total_viajes if total_viajes > 0 else 0
    eficiencia = (total_km / total_gasolina) if total_gasolina > 0 else 0

    estadisticas = {
        'total_km': round(total_km, 2),
        'total_gasolina': round(total_gasolina, 2),
        'promedio_duracion': round(promedio_duracion, 2),
        'total_viajes': total_viajes,
        'eficiencia': round(eficiencia, 2)
    }

    Itadori = {
        'cliente': user,
        'viajes_info': viajes_info,
        'estadisticas': estadisticas,
        'alertas': alertas,
        'page_title': f'Oseed | {user.conductor} {user.placa}'
    }
    
    return render(request, 'vehiculos/detallesVehiculos.html', Itadori)

@login_required
def crearVehiculos(request):
    if request.method == 'POST':
        form = VehiculoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vehiculo creado con éxito.')
            return redirect('vista_vehiculos')
        else:
            messages.error(request, 'Error al crear el vehiculo. Por favor, corrige los errores.')
    else:
        form = VehiculoForm()

    Nanami = {
        'form': form,
        'page_title': 'Oseed | Crear Vehiculo'
    }

    return render(request, 'vehiculos/crear_vehiculos.html', Nanami)

@login_required
def update_vehiculo(request, user_id):
    user = Camiones.objects.get(pk=user_id)

    if request.method == 'POST':
        form = VehiculoForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vehiculo actualizado con éxito.')
            return redirect('vista_vehiculos')
        else:
            messages.error(request, 'Error al actualizar el vehiculo. Por favor, corrige los errores.')
    else:
        form = VehiculoForm(instance=user)
        
    Nobara = {
        'form': form, 
        'user': user,
        'page_title': f'Oseed | {user.conductor} {user.placa}'
    }

    return render(request, 'vehiculos/actualizar_vehiculos.html', Nobara)

@login_required
def delete_vehiculo(request, user_id):
    user = get_object_or_404(Camiones, id=user_id)
    user.delete()
    messages.success(request, 'Vehiculo eliminado con éxito.')
    return redirect('vista_vehiculos')

@login_required
def exportar_historial_excel(request, user_id):
    user = get_object_or_404(Camiones, pk=user_id)
    viajes = UbicacionCamion.objects.filter(camion=user, viaje_completado=True).order_by('-fecha_viaje')

    def geocode(lat, lng):
        try:
            url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lng}&zoom=16&addressdetails=1"
            resp = requests.get(url, headers={'User-Agent': 'OseedApp'})
            if resp.status_code == 200:
                data = resp.json()
                return data.get('display_name', '')
        except Exception:
            return ""
        return ""

    data = []
    for v in viajes:
        inicio = f"{v.ruta[0]['lat']}, {v.ruta[0]['lng']}" if v.ruta else ""
        fin = f"{v.ruta[-1]['lat']}, {v.ruta[-1]['lng']}" if v.ruta else ""
        direccion_inicio = geocode(v.ruta[0]['lat'], v.ruta[0]['lng']) if v.ruta else ""
        direccion_fin = geocode(v.ruta[-1]['lat'], v.ruta[-1]['lng']) if v.ruta else ""
        data.append({
            'Fecha': v.fecha_viaje.strftime("%d/%m/%Y %H:%M"),
            'Inicio (Coordenadas)': inicio,
            'Dirección Inicio': direccion_inicio,
            'Fin (Coordenadas)': fin,
            'Dirección Fin': direccion_fin,
            'Puntos': len(v.ruta) if v.ruta else 0,
            'Gasolina Inicio': v.gasolina_inicio,
            'Gasolina Fin': v.gasolina_fin,
            '% Ahorrado': v.porcentaje_ahorrado,
            'Duración (min)': v.duracion_viaje,
        })

    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=historial_rutas_{user.placa}.xlsx'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        ws = writer.book.active
        # Ajustar el ancho de las columnas automáticamente
        for i, col in enumerate(df.columns, 1):
            max_length = max(
                [len(str(cell)) for cell in [col] + df[col].astype(str).tolist()]
            )
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    return response

@login_required
def obtener_ruta_viaje(request, viaje_id):
    viaje = get_object_or_404(UbicacionCamion, pk=viaje_id)
    return JsonResponse({'ruta': viaje.ruta})

@login_required
def ver_mapa_ruta(request, viaje_id):
    viaje = get_object_or_404(UbicacionCamion, pk=viaje_id)
    ruta = viaje.ruta if viaje.ruta else []
    return render(request, 'rutas_gps/mapa_ruta.html', {
        'viaje': viaje,
        'ruta': ruta,
        'page_title': f'Ruta del viaje {viaje.id}'
    })